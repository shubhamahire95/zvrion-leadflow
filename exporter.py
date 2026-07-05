"""Export clean leads to branded CSV, XLSX, and optional JSON files."""
import csv, json, re
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from config import OUTPUT_COLUMNS, OUTPUT_DIR


def _slug(value):
    return re.sub(r"[^a-z0-9]+", "_", str(value).casefold()).strip("_") or "all"


def export_leads(rows: list[dict], niche: str, location: str, output_dir: Path = OUTPUT_DIR,
                 include_json: bool = False) -> dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    stem = f"zvrion_leads_{_slug(niche)}_{_slug(location)}_{datetime.now():%Y%m%d_%H%M%S}"
    paths = {"csv": output_dir / f"{stem}.csv", "xlsx": output_dir / f"{stem}.xlsx"}
    with paths["csv"].open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS, extrasaction="ignore")
        writer.writeheader(); writer.writerows(rows)
    book = Workbook(); sheet = book.active; sheet.title = "ZVRION Leads"; sheet.append(OUTPUT_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="126BFF")
    for row in rows: sheet.append([row.get(column, "") for column in OUTPUT_COLUMNS])
    sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(c.value or "")) for c in column) + 2, 55)
    book.save(paths["xlsx"])
    if include_json:
        paths["json"] = output_dir / f"{stem}.json"
        paths["json"].write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    return paths


def export_debug(raw_rows: list[dict], rejected_rows: list[dict], output_dir: Path = OUTPUT_DIR) -> dict[str, Path]:
    """Persist source and rejected records so an empty clean export is diagnosable."""
    output_dir.mkdir(parents=True, exist_ok=True); stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    paths = {"debug_raw": output_dir / f"debug_raw_{stamp}.json"}
    paths["debug_raw"].write_text(json.dumps(raw_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    if rejected_rows:
        paths["rejected"] = output_dir / f"rejected_leads_{stamp}.json"
        paths["rejected"].write_text(json.dumps(rejected_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    return paths
